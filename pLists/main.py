



from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
print(wb.sheetnames)
wb.active = 0
ws = wb.active
print(ws['B4'].value)
ws['B4'].value = "Шанхай"
print(ws['B4'].value)
wb.save('test.xlsx')
